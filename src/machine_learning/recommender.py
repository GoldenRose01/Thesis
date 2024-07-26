#from src.machine_learning.utils import *
#from src.machine_learning.apriori import *
#from src.machine_learning.encoding import *
import settings
from src.machine_learning.filter_attributes import *
from src.machine_learning.labeling import generate_label
from src.machine_learning import evaluateEditDistance
from run_experiments import *
from src.machine_learning.decision_tree import *
from src.constants import *
from sklearn import metrics
from settings import *
import openpyxl
import pm4py
import copy
import csv

# Crea una classe ParamsOptimizer per ottimizzare i parametri del modello.
class ParamsOptimizer:
    def __init__(self, data_log, train_val_log, val_log, train_log, parameters,
                 labeling, checkers, rules, min_prefix_length, max_prefix_length):
        # Inizializza la classe con i parametri necessari.
        self.parameter_names = parameters.keys()
        # Ottiene i nomi dei parametri.
        self.val_log = val_log
        # Registro di validazione.
        self.data_log = data_log
        # Registro dati.
        self.train_val_log = train_val_log
        # Registro di allenamento e validazione.
        self.param_grid = [element for element in itertools.product(*parameters.values())]
        # Crea una griglia dei parametri possibili.
        self.train_log = train_log
        # Registro di allenamento.
        self.parameters = parameters
        # Parametri del modello.
        self.labeling = labeling
        # Etichettatura dei dati.
        self.checkers = checkers
        # Controllori.
        self.rules = rules
        # Regole.
        self.model_grid = []
        # Griglia dei modelli.
        self.min_prefix_length = min_prefix_length
        # Lunghezza minima del prefisso.
        self.max_prefix_length = max_prefix_length
        # Lunghezza massima del prefisso.

    # Funzione per la ricerca nella griglia dei parametri.
    def params_grid_search(self, dataset_name, constr_family):
        categories = [TraceLabel.FALSE.value, TraceLabel.TRUE.value]  # Definisci categorie.

        for param_id, param_tuple in enumerate(self.param_grid):
            model_dict = {'dataset_name': dataset_name, 'constr_family': constr_family, 'parameters': param_tuple,
                          'f1_score_val': None, 'f1_score_train': None, 'f1_prefix_val': None, 'max_depth': 0,
                          'id': param_id, 'model': None, 'frequent_events': None, 'frequent_pairs': None}

            (frequent_events_train, frequent_pairs_train) = generate_frequent_events_and_pairs(self.data_log,
                                                                                               param_tuple[0])

            # Genera input per l'albero decisionale
            dt_input_train = Encoding(self.train_log, labeling=self.labeling)
            dt_input_train.encode_traces
            dt_input_val = Encoding(self.val_log, labeling=self.labeling)
            dt_input_val.encode_traces

            X_train = pd.DataFrame(dt_input_train.encoded_data, columns=dt_input_train.features)
            y_train = pd.Categorical(dt_input_train.labels, categories=categories)

            X_val = pd.DataFrame(dt_input_val.encoded_data, columns=dt_input_val.features)
            y_val = pd.Categorical(dt_input_val.labels, categories=categories)

            # Genera l'albero decisionale e il punteggio sul set di validazione
            dtc, f1_score_val, f1_score_train = generate_decision_tree(X_train, X_val, y_train, y_val,
                                                                       class_weight=param_tuple[1],
                                                                       min_samples_split=param_tuple[2])
            paths = generate_paths(dtc=dtc,
                                   dt_input_features=dt_input_train.features,
                                   target_label=self.labeling["target"])

            # Valutazione sui prefissi del set di validazione
            results = []
            for pref_id, prefix_len in enumerate(range(self.min_prefix_length, self.max_prefix_length + 1)):
                prefixing = {
                    "type": PrefixType.ONLY,
                    "length": prefix_len
                }

                evaluation = evaluate_recommendations(input_log=self.val_log,
                                                      labeling=self.labeling,
                                                      prefixing=prefixing,
                                                      rules=self.rules,
                                                      paths=paths,
                                                      log=self.train_log,
                                                      dataset_name=dataset_name,
                                                      features=features,
                                                      resource_attributes=resource_attributes,
                                                      trace_attributes=trace_attributes
                                                      )
                results.append(evaluation)

            model_dict['model'] = dtc
            model_dict['f1_score_val'] = f1_score_val
            model_dict['f1_score_train'] = f1_score_train
            model_dict['f1_prefix_val'] = np.average([res.fscore for res in results])
            model_dict['frequent_events'] = frequent_events_train
            model_dict['frequent_pairs'] = frequent_pairs_train
            self.model_grid.append(model_dict)

        # Retraining dell'albero decisionale utilizzando il set di allenamento e validazione con i migliori parametri testati sul set di validazione
        sorted_models = sorted(self.model_grid, key=lambda d: d['f1_prefix_val'])
        best_model_dict = sorted_models[-1]

        # Input per l'albero decisionale con il set di allenamento e validazione
        dt_input_trainval = Encoding(self.train_val_log, labeling=self.labeling)
        dt_input_trainval.encode_traces
        dt_input_val = Encoding(self.val_log, labeling=self.labeling)
        dt_input_val.encode_traces

        X_train_val = pd.DataFrame(dt_input_trainval.encoded_data, columns=dt_input_trainval.features)
        y_train_val = pd.Categorical(dt_input_trainval.labels, categories=categories)
        X_val = pd.DataFrame(dt_input_val.encoded_data, columns=dt_input_val.features)
        y_val = pd.Categorical(dt_input_val.labels, categories=categories)

        dtc, _, _ = generate_decision_tree(X_train_val, X_val, y_train_val, y_val,
                                           class_weight=best_model_dict['parameters'][1],
                                           min_samples_split=best_model_dict['parameters'][2])
        best_model_dict['model'] = dtc
        best_model_dict['max_depth'] = dtc.tree_.max_depth

        del best_model_dict["frequent_events"]
        del best_model_dict["frequent_pairs"]
        return best_model_dict, dt_input_trainval.features

# Definisci una funzione per generare raccomandazioni basate su un prefisso e un percorso
def recommend(prefix, path, dt_input_trainval, dataset_name, features, resource_attributes, trace_attributes):
    recommendation = ""


    prefixes = []
    for trace in prefix:
        prefixes.append(trace['concept:name'])
    num_prefixes = len(prefixes)

    for rule in path.rules:
        feature, state, parent = rule

        numbers = extract_numbers_from_string(feature, trace_attributes, resource_attributes)
        if numbers is None:
            continue  # Salta questo ciclo se numbers è None

        for num_tuple in numbers:
            if len(num_tuple) == 2:
                num1, num2 = num_tuple
            elif len(num_tuple) == 1:
                num1 = num_tuple[0]
                num2 = None

            feature_index = find_feature_position(feature, features)

            if num1 == "Skip" or num1 > num_prefixes:
                rec = np.zeros(len(features), dtype=int)
                if num2 is not None:
                    rec[feature_index] = int(num2)
                rec = rec.tolist()

                rec_str = dt_input_trainval.decode(rec,features)
                for column in rec_str.columns:
                    if rec_str[column].iloc[0] != '0' and rec_str[column].notnull().any():
                        if state == TraceState.VIOLATED:
                            recommendation += f"{column} should not be {str(rec_str[column].iloc[0])}; "
                        elif state == TraceState.SATISFIED:
                            recommendation += f"{column} should be {str(rec_str[column].iloc[0])}; "

    return recommendation

def simple_extraction(trace):
    activities = []
    for idx, event in enumerate(trace):
        for attribute_key, attribute_value in event.items():
            if attribute_key == 'concept:name':
                activities.append(attribute_value)
    return activities

def complex_extraction(trace, trace_att_d, resource_att_d, prefix_max):
    output = {}

    # Initialize output with trace attributes from the first event
    first_event = trace[0] if trace else {}
    for att in trace_att_d:
        output[att] = first_event.get(att, None)

    # Initialize lists for prefixes and resources
    for i in range(prefix_max):
        output[f'prefix_{i + 1}'] = None
        for res in resource_att_d:
            output[f'{res}_{i + 1}'] = None

    # Populate output and activities
    for idx, event in enumerate(trace):
        if idx >= prefix_max:
            break

        # Populate activities
        output[f'prefix_{idx + 1}'] = event.get('concept:name', None)

        # Populate resources
        for res in resource_att_d:
            output[f'{res}_{idx + 1}'] = event.get(res, None)

    # Convert output to an ordered list of values
    ordered_output = [output[att] for att in trace_att_d]
    ordered_output += [output[f'prefix_{i + 1}'] for i in range(prefix_max)]
    for res in resource_att_d:
        ordered_output += [output[f'{res}_{i + 1}'] for i in range(prefix_max)]

    return ordered_output


# Definisci una funzione per valutare la conformità di un trace rispetto a un percorso
def evaluate(trace, path, num_prefixes, dt_input_trainval, sat_threshold,
             labeling, indices, max_variation, dataset_name, prefix_max,
             features, resource_attributes, trace_attributes):
    # Sistemazione indici
    list_of_index = array_index(prefix_max, trace_attributes, resource_attributes)
    # Creazione mappa di encoding
    encoding_map = {}

    length_t = len(trace_attributes)
    length_p = length_t + prefix_max
    length_r = length_p + (prefix_max * len(resource_attributes))

    activities = []  # Definizione della lista che conterrà gli elementi da ritornare

    if settings.type_encoding == "complex":
        activities = complex_extraction(trace, trace_attributes, resource_attributes, prefix_max)
    elif settings.type_encoding == "simple":
        activities = simple_extraction(trace)
    # elif settings.type_encoding == "frequency":
    # todo implementa encoding per frequency

    # Continue with encoding and further processing...
    activities = dt_input_trainval.encode(activities,features)
    # Pre-elaborazione di hyp da dati codificati
    hyp_t = [int(value) if isinstance(value, str) and value.isdigit() else value for value in activities.iloc[0].values]
    if settings.type_encoding == "complex":
        hyp = rm_vect_element(hyp=hyp_t,
                              list_of_index=list_of_index,
                              m=num_prefixes,
                              resource_att_d=resource_attributes)
    elif settings.type_encoding == "simple":
        hyp = hyp_t[num_prefixes:]

    n_max = 0
    # Inizializza ref con il valore massimo derivato dalle regole
    for rule in path.rules:
        feature, state, parent = rule

        numbers = extract_numbers_from_string(feature, trace_attributes, resource_attributes)
        if numbers is None:
            continue
        for num_tuple in numbers:
            if len(num_tuple) == 2:
                num1, num2 = num_tuple
            elif len(num_tuple) == 1:
                num1 = num_tuple[0]
                num2 = None

            n_max = len(features)

    ref = np.zeros(n_max, dtype=int)

    for rule in path.rules:
        feature, state, parent = rule
        numbers = extract_numbers_from_string(feature,  trace_attributes, resource_attributes)
        if numbers is None:
            continue
        for num_tuple in numbers:
            if len(num_tuple) == 2:
                num1, num2 = num_tuple
            elif len(num_tuple) == 1:
                num1 = num_tuple[0]
                num2 = None

            feature_index = find_feature_position(feature, features)

            if num1 == "Skip" or num1 > num_prefixes:
                if num2 is not None:
                    ref[feature_index] = int(num2)

    if settings.type_encoding == "complex":
        ref = rm_vect_element(hyp=ref, list_of_index=list_of_index,
                              m=num_prefixes, resource_att_d=resource_attributes)
    elif settings.type_encoding == "simple":
        ref = ref[num_prefixes:]

    ref = ref.tolist()

    ed = 0
    # Override di sicurezza per simple encoding
    if settings.type_encoding == "simple":
        ed = evaluateEditDistance.edit(ref, hyp)
    elif settings.type_encoding == "complex":
        if selected_evaluation_edit_distance == "edit":
            # Calcolo della distanza pura basata sulla libreria editdistance
            ed = evaluateEditDistance.edit(ref, hyp)
        elif selected_evaluation_edit_distance == "edit_separate":
            # Calcolo della distanza pura basata su num e categoric
            ed = evaluateEditDistance.edit_separate(ref, hyp, indices, max_variation)
        elif selected_evaluation_edit_distance == "weighted_edit_distance":
            ed = evaluateEditDistance.weighted_edit_distance(ref, hyp, indices, max_variation, length_t)

    if ed < sat_threshold:
        is_compliant = True
    else:
        is_compliant = False

    label = generate_label(trace, labeling)

    if is_compliant:
        cm = ConfusionMatrix.TP if label == TraceLabel.TRUE else ConfusionMatrix.FP
    else:
        cm = ConfusionMatrix.FN if label == TraceLabel.TRUE else ConfusionMatrix.TN

    return is_compliant, cm

# Definisci una funzione per testare il decision tree
def test_dt(test_log, train_log, labeling, prefixing, support_threshold, checkers, rules):

    (frequent_events, frequent_pairs) = generate_frequent_events_and_pairs(train_log, support_threshold)

    print("Generazione dell'input per il decision tree ...")
    dt_input = Encoding(train_log, labeling)
    dt_input.encode_traces

    print("Generazione del decision tree ...")
    return dt_score(dt_input=dt_input)


def test_dt(test_log, train_log, labeling, prefixing, support_threshold, checkers, rules):
    (frequent_events, frequent_pairs) = generate_frequent_events_and_pairs(train_log, support_threshold)

    print("Generating decision tree input ...")
    dt_input = Encoding(train_log, labeling)
    dt_input.encode_traces

    print("Generating decision tree ...")
    return dt_score(dt_input=dt_input)


def train_path_recommender(data_log, train_val_log, val_log, train_log, labeling, support_threshold,
                           dataset_name, output_dir, dt_input_trainval, resource_attributes, trace_attributes):

    if labeling["threshold_type"] == LabelThresholdType.LABEL_MEAN:
        labeling["custom_threshold"] = calc_mean_label_threshold(train_log, labeling)

    target_label = labeling["target"]

    at_gendt=f"Generating decision tree with params optimization"
    print(f"{JUNGLE_GREEN}{at_gendt.center(main.infoconsole())}{RESET}")

    if settings.optimize_dt:
        best_model_dict, feature_names = find_best_dt(dataset_name, train_val_log,
                                                      support_threshold, settings.print_dt, dt_input_trainval,
                                                      resource_attributes, trace_attributes)
    # else:
    # param_opt = ParamsOptimizer(data_log, train_val_log, val_log, train_log, settings.hyperparameters, labeling,
    #                            checkers, rules, min_prefix_length, max_prefix_length)
    # best_model_dict, feature_names = param_opt.params_grid_search(dataset_name, constr_family)

    encoding = f"{settings.type_encoding} with {settings.selected_evaluation_edit_distance}"
    if settings.selected_evaluation_edit_distance == "weighted_edit_distance":
        proportion = f"{settings.wtrace_att},{settings.wactivities},{settings.wresource_att}"
    else:
        proportion = "Null"

    with open(os.path.join(output_dir, 'model_params.csv'), 'a', newline='') as f:
        w = csv.writer(f, delimiter='\t')
        row = list(best_model_dict.values())
        row.insert(0, encoding)  # Aggiungi Encoding all'inizio della riga
        row.insert(1, proportion)  # Aggiungi Proportion dopo Encoding
        w.writerow(row[:-1])

    if settings.Allprint:
        at_tp= f"Generating decision tree paths"
        print(at_tp.center(main.infoconsole()))

    paths = generate_paths(dtc=best_model_dict['model'], dt_input_features=feature_names,
                           target_label=target_label)
    return paths, best_model_dict

def evaluate_recommendations(input_log, labeling, prefixing, rules, paths, train_log,dataset_name,features,
                             resource_attributes, trace_attributes):
    # if labeling["threshold_type"] == LabelThresholdType.LABEL_MEAN:
    #    labeling["custom_threshold"] = calc_mean_label_threshold(train_log, labeling)
    features = features[1:]
    target_label = labeling["target"]
    if settings.allprint==True:
        at_gentest=f"Generating test prefixes"
        print(at_gentest.center(main.infoconsole()))
    prefixes = generate_prefixes(input_log, prefixing)

    eval_res = EvaluationResult()

    for prefix_length in prefixes:
        # for id, pref in enumerate(prefixes[prefix_length]): print(id, input_log[pref.trace_num][0]['label'])
        for prefix in prefixes[prefix_length]:
            for path in paths:
                path.fitness = calcPathFitnessOnPrefix(prefix.events, path, rules, settings.fitness_type,train_log,
                                                       features, resource_attributes, trace_attributes)

            paths = sorted(paths, key=lambda path: (- path.fitness, path.impurity, - path.num_samples["total"]),
                           reverse=False)

            selected_path = None

            for path_index, path in enumerate(paths):

                if selected_path and (
                        path.fitness != selected_path.fitness
                        or path.impurity != selected_path.impurity
                        or path.num_samples != selected_path.num_samples):
                    break

                recommendation = recommend(prefix.events, path, rules,dataset_name,features,
                                           resource_attributes, trace_attributes)
                # print(f"{prefix_length} {prefix.trace_num} {prefix.trace_id} {path_index}->{recommendation}")
                trace = input_log[prefix.trace_num]

                if recommendation != "Contradiction" and recommendation != "":
                    # if recommendation != "":
                    selected_path = path
                    trace = input_log[prefix.trace_num]
                    is_compliant, e = evaluate(trace, path, rules, labeling,features,
                                               resource_attributes, trace_attributes)

                    """
                    if prefix_length > 2:
                        # for event in prefix.events: print(event['concept:name'])
                        # for path in paths: print(path.rules)
                        #recommend(prefix.events, paths[0], rules)
                        #len(paths[0])
                        #for rule in paths[0]: print(rule)
                        for path in paths: print(evaluate(trace, path, rules, labeling))
                    """

                    if e == ConfusionMatrix.TP:
                        eval_res.tp += 1
                    elif e == ConfusionMatrix.FP:
                        eval_res.fp += 1
                    elif e == ConfusionMatrix.FN:
                        eval_res.fn += 1
                    elif e == ConfusionMatrix.TN:
                        eval_res.tn += 1

    try:
        eval_res.precision = eval_res.tp / (eval_res.tp + eval_res.fp)
    except ZeroDivisionError:
        eval_res.precision = 0

    try:
        eval_res.recall = eval_res.tp / (eval_res.tp + eval_res.fn)
    except ZeroDivisionError:
        eval_res.recall = 0

    try:
        eval_res.fscore = 2 * eval_res.precision * eval_res.recall / (eval_res.precision + eval_res.recall)
    except ZeroDivisionError:
        eval_res.fscore = 0

    return eval_res


def generate_recommendations_and_evaluation(test_log,
                                            train_log,
                                            labeling,
                                            prefixing,
                                            rules,
                                            paths,
                                            hyperparams_evaluation,
                                            dt_input_trainval,
                                            dt_input_trainval_encoded,
                                            indices,
                                            max_variations,
                                            dataset_name,
                                            prefix_max,
                                            features,
                                            resource_attributes,
                                            trace_attributes,
                                            eval_res=None,
                                            debug=False):
    features = features[1:]
    if labeling["threshold_type"] == LabelThresholdType.LABEL_MEAN:
        labeling["custom_threshold"] = calc_mean_label_threshold(train_log, labeling)

    target_label = labeling["target"]

    if print_log==True:
        print(test_log)
    if settings.Allprint is True:
        print("Generating test prefixes ...")

    test_prefixes = generate_prefixes(test_log, prefixing)

    if print_length==True:
        print(prefixing["length"])
    if settings.Allprint==True:
        print("Generating recommendations ...")
    recommendations = []
    if eval_res is None:
        eval_res = EvaluationResult()
    y = []
    pred = []

    for prefix_length in test_prefixes:
        eval_res.prefix_length = prefix_length
        # for id, pref in enumerate(test_prefixes[prefix_length]): print(id, test_log[pref.trace_num][0]['label'])
        for prefix in test_prefixes[prefix_length]:
            eval_res.num_cases = len(test_prefixes[prefix_length])
            pos_paths_total_samples = 0
            for path in paths:
                pos_paths_total_samples += path.num_samples['node_samples']
            for path in paths:
                path.fitness = calcPathFitnessOnPrefix(prefix.events, path, dt_input_trainval,
                                                       train_log,dataset_name,features,
                                                       resource_attributes, trace_attributes)
                path.score = calcScore(path, pos_paths_total_samples, weights=hyperparams_evaluation[1:])
            # paths = sorted(paths, key=lambda path: (- path.fitness, path.impurity,
            #                                            - path.num_samples["total"]), reverse=False)
            if settings.use_score:
                paths = sorted(paths, key=lambda path: (- path.score), reverse=False)
            else:
                paths = sorted(paths, key=lambda path: (- path.fitness), reverse=False)
            reranked_paths = copy.deepcopy(paths)

            if settings.reranking:  # è false
                reranked_paths = paths[:settings.top_K_paths]
                reranked_paths = sorted(reranked_paths, key=lambda path: (- path.score), reverse=False)

            if settings.compute_gain and len(reranked_paths) > 0:  # è false
                raw_prefix = [event['concept:name'] for event in prefix.events]
                trace = test_log[prefix.trace_num]
                path = reranked_paths[0]
                label = generate_label(trace, labeling)
                compliant, _ = evaluate(trace, path, rules, labeling,
                                        eval_type=settings.sat_type,prefix_max=prefix_max,
                                        resource_attributes=resource_attributes, trace_attributes=trace_attributes)
                eval_res.comp += 1 if compliant else 0
                eval_res.non_comp += 0 if compliant else 1
                eval_res.pos_comp += 1 if compliant and label.value == TraceLabel.TRUE.value else 0
                eval_res.pos_non_comp += 1 if not compliant and label.value == TraceLabel.TRUE.value else 0

            selected_path = None
            for path_index, path in enumerate(reranked_paths):
                if selected_path and (path.fitness != selected_path.fitness or path.impurity != selected_path.impurity
                                      or path.num_samples != selected_path.num_samples):
                    break

                recommendation = recommend(prefix.events, path, dt_input_trainval, dataset_name,
                                           features, resource_attributes, trace_attributes)
                if settings.Allprint is True:
                    print(recommendation)
                    print(f"Prefix Lenght:{prefix_length}, {prefix.trace_num}"
                          f" {prefix.trace_id} {path_index}->{recommendation}")

                if recommendation != "Contradiction" and recommendation != "":
                    # if True:
                    # if recommendation != "":

                    selected_path = path
                    trace = test_log[prefix.trace_num]
                    # print(prefix.trace_id, trace[0]['label'])
                    is_compliant, e = evaluate(trace=trace,
                                               path=path,
                                               num_prefixes=prefix_length,
                                               dt_input_trainval=dt_input_trainval,
                                               sat_threshold=hyperparams_evaluation[0],
                                               labeling=labeling,
                                               indices=indices,
                                               max_variation=max_variations,
                                               dataset_name=dataset_name,
                                               prefix_max=prefix_max,
                                               features=features,
                                               resource_attributes=resource_attributes,
                                               trace_attributes=trace_attributes
                                               )
                    if e is None:
                        e_name = "UnknownError"
                    else:
                        e_name = e.name

                    # if prefix_length == 12 or prefix_length == 12:
                    # pdb.set_trace()
                    # pdb.set_trace()
                    if debug:
                        # pdb.set_trace()
                        if prefix.trace_id == 'GX' and prefix_length == 15:
                            for event in prefix.events: print(event['concept:name'])
                            # pdb.set_trace()
                        if prefix.trace_id == 'DS' and prefix_length == 5:
                            for event in prefix.events: print(event['concept:name'])
                            print(e_name)
                            print(prefix_length)
                            # pdb.set_trace()

                    if e == ConfusionMatrix.TP:
                        eval_res.tp += 1
                    elif e == ConfusionMatrix.FP:
                        eval_res.fp += 1
                    elif e == ConfusionMatrix.FN:
                        eval_res.fn += 1
                    elif e == ConfusionMatrix.TN:
                        eval_res.tn += 1



                    prefixes = []
                    for n in prefix.events:
                        prefixes.append(n['concept:name'])

                    recommendation_model = Recommendation(
                        trace_id=prefix.trace_id,
                        prefix_len=len(prefixes),
                        complete_trace=generate_prefix_path(test_log[prefix.trace_num]),
                        current_prefix=generate_prefix_path(prefix.events),
                        actual_label=generate_label(trace, labeling).name,
                        target_label=target_label.name,
                        is_compliant=str(is_compliant).upper(),
                        confusion_matrix=e_name,
                        impurity=path.impurity,
                        num_samples=path.num_samples,
                        fitness=path.fitness,
                        score=path.score,
                        recommendation=recommendation
                    )
                    y.append(recommendation_model.actual_label)
                    pred.append(
                        recommendation_model.num_samples["positive"] / recommendation_model.num_samples["total"])
                    # pred.append(recommendation_model.score)
                    recommendations.append(recommendation_model)

    try:
        eval_res.precision = eval_res.tp / (eval_res.tp + eval_res.fp)
    except ZeroDivisionError:
        eval_res.precision = 0

    try:
        eval_res.recall = eval_res.tp / (eval_res.tp + eval_res.fn)
    except ZeroDivisionError:
        eval_res.recall = 0

    try:
        eval_res.accuracy = (eval_res.tp + eval_res.tn) / (eval_res.tp + eval_res.fp + eval_res.fn + eval_res.tn)
    except ZeroDivisionError:
        eval_res.accuracy = 0

    try:
        eval_res.fscore = 2 * eval_res.precision * eval_res.recall / (eval_res.precision + eval_res.recall)
    except ZeroDivisionError:
        eval_res.fscore = 0
    try:
        fpr, tpr, thresholds = metrics.roc_curve(np.array(y), np.array(pred), pos_label=target_label.name)
        eval_res.auc = metrics.auc(fpr, tpr)
    except:
        eval_res.auc = 0

    try:
        eval_res.mcc = matthews_corrcoef(eval_res.tp, eval_res.fp, eval_res.fn, eval_res.tn)
    except:
        eval_res.mcc = 0

    if settings.compute_gain:
        eval_res.gain = gain(eval_res.comp, eval_res.non_comp, eval_res.pos_comp, eval_res.pos_non_comp)

    if settings.eval_stamp is True:
        print("Writing evaluation result into csv file ...")
        write_evaluation_to_csv(eval_res, dataset_name)

    if settings.recc_stamp is True:
        print("Writing recommendations into csv file ...")
        write_recommendations_to_csv(recommendations, dataset_name)

    return recommendations, eval_res


def write_evaluation_to_csv(e, dataset):
    if selected_evaluation_edit_distance != "weighted_edit_distance":
        csv_file = os.path.join(settings.results_dir, f"{dataset}_{ruleprefix}{type_encoding}_"
                                                      f"evaluation_{selected_evaluation_edit_distance}.csv")
    else:
        csv_file = os.path.join(settings.results_dir,
                                f"{dataset}_{ruleprefix}{type_encoding}_evaluation_"
                                f"{selected_evaluation_edit_distance}{settings.wtrace_att},"
                                f"{settings.wactivities},{settings.wresource_att}.csv")
    fieldnames = ["tp", "fp", "tn", "fn", "precision", "recall", "accuracy", "fscore", "auc"]
    values = {
        "tp": e.tp,
        "fp": e.fp,
        "tn": e.tn,
        "fn": e.fn,
        "precision": round(e.precision, 2),
        "recall": round(e.recall, 2),
        "accuracy": round(e.accuracy, 2),
        "fscore": round(e.fscore, 2),
        "auc": round(e.auc, 2)
    }
    try:
        with open(csv_file, 'w') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerow(values)
    except IOError:
        print("I/O error")


def write_recommendations_to_csv(recommendations, dataset):
    if selected_evaluation_edit_distance != "weighted_edit_distance":
        csv_file = os.path.join(settings.results_dir,
                                f"{dataset}_{ruleprefix}{type_encoding}_recommendations_"
                                f"{selected_evaluation_edit_distance}.csv")
    else:
        csv_file = os.path.join(settings.results_dir,
                                f"{dataset}_{ruleprefix}{type_encoding}_recommendations_"
                                f"{selected_evaluation_edit_distance}{settings.wtrace_att},"
                                f"{settings.wactivities},{settings.wresource_att}.csv")

    fieldnames = ["Trace id", "Prefix len", "Complete trace", "Current prefix", "Recommendation", "Actual label",
                  "Target label", "Compliant", "Confusion matrix", "Impurity", "Fitness", "Num samples"]

    # Write to CSV
    values = []
    for r in recommendations:
        values.append(
            {
                "Trace id": r.trace_id,
                "Prefix len": r.prefix_len,
                "Complete trace": r.complete_trace,
                "Current prefix": r.current_prefix,
                "Recommendation": r.recommendation,
                "Actual label": r.actual_label,
                "Target label": r.target_label,
                "Compliant": r.is_compliant,
                "Confusion matrix": r.confusion_matrix,
                "Impurity": r.impurity,
                "Fitness": r.fitness,
                "Num samples": r.num_samples["total"]
            }
        )

    try:
        with open(csv_file, 'w') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for value in values:
                writer.writerow(value)
    except IOError:
        print("I/O error")

    # Write to Excel
    if selected_evaluation_edit_distance != "weighted_edit_distance":
        excel_file = os.path.join(settings.results_dir,
                                  f"{dataset}_{ruleprefix}{type_encoding}_recommendations_"
                                  f"{selected_evaluation_edit_distance}.xlsx")
    else:
        excel_file = os.path.join(settings.results_dir,
                                  f"{dataset}_{ruleprefix}{type_encoding}_recommendations_"
                                  f"{selected_evaluation_edit_distance}{settings.wtrace_att},"
                                  f"{settings.wactivities},{settings.wresource_att}.xlsx")
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Recommendations"

    # Write headers
    for col_idx, field in enumerate(fieldnames, start=1):
        worksheet.cell(row=1, column=col_idx).value = field

    # Write data
    for row_idx, value in enumerate(values, start=2):
        for col_idx, field_value in enumerate(value.values(), start=1):
            worksheet.cell(row=row_idx, column=col_idx).value = field_value

    workbook.save(excel_file)


def prefix_evaluation_to_csv(result_dict, dataset):
    if selected_evaluation_edit_distance != "weighted_edit_distance":
        csv_file = os.path.join(settings.results_dir,
                                f"{dataset}_{ruleprefix}{type_encoding}_evaluation_"
                                f"{selected_evaluation_edit_distance}.csv")
        excel_file = os.path.join(settings.results_dir,
                                  f"{dataset}_{ruleprefix}{type_encoding}_evaluation_"
                                  f"{selected_evaluation_edit_distance}.xlsx")
    else:
        csv_file = os.path.join(settings.results_dir,
                                f"{dataset}_{ruleprefix}{type_encoding}_evaluation_"
                                f"{selected_evaluation_edit_distance}{settings.wtrace_att},"
                                f"{settings.wactivities},{settings.wresource_att}.csv")
        excel_file = os.path.join(settings.results_dir,
                                  f"{dataset}_{ruleprefix}{type_encoding}_evaluation_"
                                  f"{selected_evaluation_edit_distance}{settings.wtrace_att},"
                                  f"{settings.wactivities},{settings.wresource_att}.xlsx")


    fieldnames = ["prefix_length", "num_cases", "tp", "fp", "tn", "fn", "precision", "recall", "fscore"]

    # Write to CSV with improved clarity
    try:
        with open(csv_file, 'w', newline='') as f:
            writer = csv.writer(f, delimiter=',')
            writer.writerow(fieldnames)  # Write header row

            # Prepare data list with concise variable names
            data_list = []
            for eval_obj in result_dict:
                data = [  # Use a shorter variable name for clarity
                    eval_obj.prefix_length,
                    eval_obj.num_cases,
                    getattr(eval_obj, "tp"),
                    getattr(eval_obj, "fp"),
                    getattr(eval_obj, "tn"),
                    getattr(eval_obj, "fn"),
                    getattr(eval_obj, "precision"),
                    getattr(eval_obj, "recall"),
                    getattr(eval_obj, "fscore"),
                ]
                data_list.append(data)

            writer.writerows(data_list)  # Write all data rows at once

    except IOError:
        print("I/O error")

    # Write to Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Evaluation"

    # Write headers
    for col_idx, field in enumerate(fieldnames, start=1):
        worksheet.cell(row=1, column=col_idx).value = field

    # Write data
    data_list = []  # Reuse the data list for Excel writing
    for eval_obj in result_dict:
        data = [
            eval_obj.prefix_length,
            eval_obj.num_cases,
            getattr(eval_obj, "tp"),
            getattr(eval_obj, "fp"),
            getattr(eval_obj, "tn"),
            getattr(eval_obj, "fn"),
            getattr(eval_obj, "precision"),
            getattr(eval_obj, "recall"),
            getattr(eval_obj, "fscore"),
        ]
        data_list.append(data)

    for row_idx, row in enumerate(data_list, start=2):
        for col_idx, value in enumerate(row, start=1):
            worksheet.cell(row=row_idx, column=col_idx).value = value

    workbook.save(excel_file)